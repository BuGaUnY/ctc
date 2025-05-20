
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from django.utils.dateparse import parse_date
from django.db.models import Count, Q , Case, When, IntegerField, F, Value
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView, View, TemplateView
from .models import Activity, Organizer , Attendance, AttendanceCheckin, Ticket
from base.models import Profile
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.messages.views import SuccessMessageMixin
from django.contrib import messages
from .forms import TicketForm, AttendanceCheckinForm, OrganizerForm, ActivityForm, ReportExportForm
from django.urls import reverse, reverse_lazy
from django_filters.views import FilterView
from django_filters import FilterSet, RangeFilter, DateRangeFilter, DateFilter, ChoiceFilter
import django_filters
from django import forms
from datetime import datetime
from django.http import HttpResponse, HttpResponseBadRequest, Http404 ,HttpResponseNotAllowed, JsonResponse
from django.utils import timezone
from django.forms.models import modelformset_factory
from django.core.exceptions import ValidationError, ObjectDoesNotExist
import logging , uuid , re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from django.views import View
from django.db import transaction
from django.db.models.functions import Cast, Concat
from django.db import models

logger = logging.getLogger(__name__)

def is_valid_uuid(value):
    try:
        uuid.UUID(value)
        return True
    except ValueError:
        return False

class Teacher(LoginRequiredMixin,ListView):
    model = Organizer
    template_name = 'activity/teacher.html'
    context_object_name = 'teachers'
    ordering = ['-date_create']
    
    def get_queryset(self):
        return Organizer.objects.filter(owner=self.request.user.profile)

class AddOrganizerView(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = Organizer
    form_class = OrganizerForm
    template_name = 'activity/add-organizer.html'
    success_url = reverse_lazy('organizer-list')
    success_message = 'เพิ่มครูที่ปรึกษา | เจ้าหน้าที่เรียบร้อยแล้ว'


class AddActivityView(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = Activity
    form_class = ActivityForm
    template_name = 'activity/add-activity.html'
    success_url = reverse_lazy('activity-list')  # Replace with the URL name you want to redirect to after saving
    success_message = 'เพิ่มกิจกรรมพิเศษเรียบร้อยแล้ว'

class ActivityUpdateView(UpdateView, SuccessMessageMixin):
    model = Activity
    form_class = ActivityForm
    template_name = 'activity/edit_activity.html'
    success_message = 'แก้ไขข้อมูลกิจกรรมพิเศษเรียบร้อยแล้ว'

    def get_success_url(self):
        return reverse_lazy('activity-list', kwargs={'pk': self.object.pk})

class DeleteActivityView(LoginRequiredMixin, SuccessMessageMixin, DeleteView):
    model = Activity
    template_name = 'activity/confirm_delete.html'  # Create this template to confirm deletion
    success_url = reverse_lazy('activity-list')  # Redirect to the activity list after deletion
    success_message = 'ลบกิจกรรมพิเศษเรียบร้อยแล้ว'

class OrganizerList(ListView):
    model = Organizer
    template_name = 'activity/organizer-list.html'
    context_object_name = 'organizers'
    ordering = ['-date_create']
    paginate_by = 9

class OrganizerDetail(DetailView):
    model = Organizer
    template_name = 'activity/organizer-detail.html'
    context_object_name = 'organizer'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['activitys'] = Activity.objects.filter(
            organizer=self.object
        ) | Activity.objects.filter(
            organizer1=self.object
        ) | Activity.objects.filter(
            organizer2=self.object
        ) | Activity.objects.filter(
            organizer3=self.object
        ) | Activity.objects.filter(
            organizer4=self.object
        ) | Activity.objects.filter(
            organizer5=self.object
        )| Activity.objects.filter(
            organizer6=self.object
        )| Activity.objects.filter(
            organizer7=self.object
        )| Activity.objects.filter(
            organizer8=self.object
        )| Activity.objects.filter(
            organizer9=self.object
        )
        return context

class OrganizerOwnerList(LoginRequiredMixin, ListView):
    model = Organizer
    template_name = 'activity/organizer-owner-list.html'
    context_object_name = 'organizers'
    ordering = ['-date_create']

    def get_queryset(self):
        return Organizer.objects.filter(owner=self.request.user.profile)

class OrganizerOwnerDetail(LoginRequiredMixin, DetailView):
    model = Organizer
    template_name = 'activity/organizer-owner-detail.html'
    context_object_name = 'organizer'

    def get_queryset(self):
        return Organizer.objects.filter(owner=self.request.user.profile, pk=self.kwargs['pk'])

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # รวมการกรอง activity จาก organizer, organizer1, และ organizer2
        context['activitys'] = Activity.objects.filter(
            organizer=self.object
        ) | Activity.objects.filter(
            organizer1=self.object
        ) | Activity.objects.filter(
            organizer2=self.object
        ) | Activity.objects.filter(
            organizer3=self.object
        ) | Activity.objects.filter(
            organizer4=self.object
        ) | Activity.objects.filter(
            organizer5=self.object
        )| Activity.objects.filter(
            organizer6=self.object
        )| Activity.objects.filter(
            organizer7=self.object
        )| Activity.objects.filter(
            organizer8=self.object
        )| Activity.objects.filter(
            organizer9=self.object
        )
        return context
    
class OrganizerOwnerActivityCheckin(LoginRequiredMixin, TemplateView):
    template_name = 'activity/organizer-owner-activity-checkin.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['activity'] = Activity.objects.get(pk=self.kwargs['activity_pk'])
        context['organizer'] = Organizer.objects.get(pk=self.kwargs['organizer_pk'])
        context['org_pk'] = self.kwargs['organizer_pk']
        return context
    
    def post(self, request, *args, **kwargs):
        org_pk = kwargs['organizer_pk']
        act_pk = kwargs['activity_pk']

        # Debug logging to check the received values
        logger.debug(f"Received org_pk: {org_pk}, act_pk: {act_pk}")

        # Validate if both primary keys are valid
        try:
            Organizer.objects.get(pk=org_pk)
            Activity.objects.get(pk=act_pk)
        except (Organizer.DoesNotExist, Activity.DoesNotExist) as e:
            logger.error(f"Invalid primary key: {str(e)}")
            return JsonResponse({'success': False, 'error': 'Invalid organizer or activity ID'}, status=400)

        url = reverse('organizer-owner-activity-checkin', kwargs={'organizer_pk': org_pk, 'activity_pk': act_pk})
        return redirect(url)

class OrganizerOwnerActivityTicketList(LoginRequiredMixin, ListView):
    model = Ticket
    template_name = 'activity/organizer-owner-activity-ticket-list.html'
    context_object_name = 'tickets'
    ordering = ['-date_create']

    def get_queryset(self):
        activity = get_object_or_404(Activity, pk=self.kwargs['pk'])
        user_profile = self.request.user.profile

        # Filter tickets by any of the organizer fields being related to the user's profile
        return Ticket.objects.filter(
            Q(activity=activity) & 
            (Q(activity__organizer__owner=user_profile) |
             Q(activity__organizer1__owner=user_profile) |
             Q(activity__organizer2__owner=user_profile) |
             Q(activity__organizer3__owner=user_profile) |
             Q(activity__organizer4__owner=user_profile) |
             Q(activity__organizer5__owner=user_profile) |
             Q(activity__organizer6__owner=user_profile) |
             Q(activity__organizer7__owner=user_profile) |
             Q(activity__organizer8__owner=user_profile) |
             Q(activity__organizer9__owner=user_profile))
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Ensure the context has the correct activity where the user is an organizer (or organizer1-9)
        context['activity'] = get_object_or_404(
            Activity,
            Q(organizer__owner=self.request.user.profile) |
            Q(organizer1__owner=self.request.user.profile) |
            Q(organizer2__owner=self.request.user.profile) |
            Q(organizer3__owner=self.request.user.profile) |
            Q(organizer4__owner=self.request.user.profile) |
            Q(organizer5__owner=self.request.user.profile) |
            Q(organizer6__owner=self.request.user.profile) |
            Q(organizer7__owner=self.request.user.profile) |
            Q(organizer8__owner=self.request.user.profile) |
            Q(organizer9__owner=self.request.user.profile),
            pk=self.kwargs['pk']  # The keyword argument is placed last
        )
        return context

class ActivityFilter(django_filters.FilterSet):
    class Meta:
        model = Activity
        fields = ['organizer', 'date_start', 'activity_category']

    def __init__ (self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.filters['organizer'].label = 'ผู้จัดงาน'
        self.filters['date_start'].label = 'ช่วงเวลาจัดงาน'
        self.filters['activity_category'].label = 'ประเภทงาน'

class ActivitySearch(FilterView):
    template_name = 'activity/activity-search.html'
    filterset_class = ActivityFilter

class ActivityList(ListView):
    model = Activity
    template_name = 'activity/activity-list.html'
    context_object_name = 'activitys'
    paginate_by = 6
    ordering = ['-date_create']

class ActivityDetail(DetailView):
    model = Activity
    template_name = 'activity/activity-detail.html'
    context_object_name = 'activity'

class TicketCreate(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = Ticket
    template_name = 'activity/ticket-create.html'
    form_class = TicketForm
    success_message = 'บันทึกข้อมูลเรียบร้อยแล้ว'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['activity'] = Activity.objects.get(pk=self.kwargs['pk'])
        return context

    def get_initial(self):
        initial = super().get_initial()
        profile = self.request.user.profile
        initial['student_number'] = profile.student_number
        initial['first_name'] = profile.first_name
        initial['last_name'] = profile.last_name
        initial['room'] = profile.room
        initial['degree'] = profile.degree
        initial['department'] = profile.department
        return initial

    def form_valid(self, form):
        form.instance.profile = self.request.user.profile
        form.instance.activity = Activity.objects.get(pk=self.kwargs['pk'])
        profile = form.instance.profile
        activity = form.instance.activity
        if Ticket.objects.filter(activity=activity, profile=profile).exists():
            messages.error(self.request, 'คุณได้ลงทะเบียนไปแล้ว')
            return super().form_invalid(form)
        return super().form_valid(form)

    def form_invalid(self, form):
        return super().form_invalid(form)
    
    def get_success_url(self):
        return reverse('ticket-list')

class TicketList(LoginRequiredMixin, ListView):
    model = Ticket
    template_name = 'activity/ticket-list.html'
    context_object_name = 'tickets'
    paginate_by = 10
    ordering = ['-date_create']

    def get_queryset(self):
        return Ticket.objects.filter(profile=self.request.user.profile)

class TicketDetail(LoginRequiredMixin, DetailView):
    model = Ticket
    template_name = 'activity/ticket-detail.html'
    context_object_name = 'ticket'

    def get_queryset(self):
        ticket = Ticket.objects.filter(profile=self.request.user.profile, pk=self.kwargs['pk'])
        return ticket

class TicketUpdate(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = Ticket
    template_name = 'activity/ticket-update.html'
    form_class = TicketForm
    success_message = 'บันทึกข้อมูลเรียบร้อยแล้ว'

    def get_queryset(self):
        ticket = Ticket.objects.filter(profile=self.request.user.profile, pk=self.kwargs['pk'])
        return ticket

    def get_success_url(self):
        return reverse('ticket-detail', kwargs={'pk': self.kwargs['pk']})

class TicketCheckin(LoginRequiredMixin, View):
    success_message = 'เช็คอินเรียบร้อยแล้ว'
    def get(self, request, *args, **kwargs):
        # ดึงค่าพารามิเตอร์จาก GET request
        activity_uid = request.GET.get('activity_uid')
        ticket_uid = request.GET.get('ticket_uid')

        # ตรวจสอบว่ามีการส่งค่าทั้งสองตัวเข้ามาหรือไม่
        if not activity_uid or not ticket_uid:
            return JsonResponse({'success': False, 'error': 'Missing activity_uid or ticket_uid'}, status=400)

        # พยายามค้นหา Activity และ Ticket
        try:
            activity = Activity.objects.get(uid=activity_uid)
            ticket = Ticket.objects.get(uid=ticket_uid, activity=activity)
        except Activity.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Activity not found'}, status=404)
        except Ticket.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Ticket not found'}, status=404)

        # ถ้าพบทั้ง Activity และ Ticket ให้ส่งข้อมูลไปยัง template
        return render(request, 'activity/partials/ticket-checkin.html', {'ticket': ticket})

class TicketCheckinSuccessView(View):
    def post(self, request, *args, **kwargs):
        ticket_uid = request.POST.get('ticket_uid')

        # ใช้ transaction.atomic เพื่อเริ่มต้นธุรกรรม
        with transaction.atomic():
            # ดึงอ็อบเจ็กต์ตั๋วหรือส่งกลับ 404 ถ้าไม่พบ
            ticket = get_object_or_404(Ticket.objects.select_for_update(), uid=ticket_uid)

            print(f"Before checkin: {ticket.checkin}")

            # ตรวจสอบสถานะการเช็คอิน
            if not ticket.checkin:
                ticket.checkin = True  # อัปเดตสถานะการเช็คอิน
                ticket.save()
                print(f"After save: {ticket.checkin}")

                # รีเฟรชอ็อบเจ็กต์ตั๋วจากฐานข้อมูลเพื่อให้แน่ใจว่าค่าได้รับการอัปเดต
                ticket.refresh_from_db()
                print(f"Refreshed checkin status after refresh: {ticket.checkin}")  # ตรวจสอบว่าอัปเดตถูกต้อง

                return render(request, 'activity/partials/ticket-checkin.html', {'ticket': ticket})
            else:
                return JsonResponse({'success': False, 'message': 'Ticket already checked in.'}, status=400)

class AttendanceList(ListView):
    model = Attendance
    template_name = 'attendance/attendance_list.html'
    context_object_name = 'attendances'

def bulk_checkin(request, pk):
    # ค้นหาบันทึกการเข้าร่วม
    attendance = get_object_or_404(Attendance, pk=pk)
    room_filter = request.GET.get('room')
    department_filter = request.GET.get('department')

    # ค้นหาข้อมูลโปรไฟล์ที่ไม่มีค่า null หรือค่าว่างใน 'room' และ 'department'
    profiles = Profile.objects.exclude(room__isnull=True, room='').exclude(department__isnull=True, department='')

    # กรองโปรไฟล์ตามห้องและแผนก
    if room_filter:
        profiles = profiles.filter(room=room_filter)

    if department_filter:
        profiles = profiles.filter(department=department_filter)

    # เรียงห้องตามลำดับตัวอักษร และตัวเลข A-Z, 1-9
    profiles = profiles.annotate(
        room_number=Cast(F('room'), output_field=models.CharField())
    ).order_by('room_number')

    # เรียงโปรไฟล์ตาม student_number
    profiles = profiles.order_by('student_number')

    # สร้าง FormSet สำหรับการบันทึก AttendanceCheckin
    AttendanceCheckinFormSet = modelformset_factory(
        AttendanceCheckin,
        form=AttendanceCheckinForm,
        extra=len(profiles)
    )

    date_checkin = timezone.now().date()  # วันที่ปัจจุบัน

    if request.method == 'POST':
        formset = AttendanceCheckinFormSet(request.POST)

        if formset.is_valid():
            instances = formset.save(commit=False)
            saved_count = 0
            skipped_count = 0

            for instance in instances:
                instance.att_name = attendance
                instance.date_checkin = date_checkin

                # ตรวจสอบข้อมูลซ้ำในวันนี้โดยใช้ student_number
                exists_by_student = AttendanceCheckin.objects.filter(
                    student_number=instance.student_number,
                    att_name=attendance,
                    date_checkin=date_checkin
                ).exists()

                # ตรวจสอบข้อมูลซ้ำโดยใช้ first_name และ last_name
                exists_by_name = AttendanceCheckin.objects.filter(
                    first_name=instance.first_name,
                    last_name=instance.last_name,
                    att_name=attendance,
                    date_checkin=date_checkin
                ).exists()

                if exists_by_student or exists_by_name:
                    skipped_count += 1
                    continue

                instance.save()
                saved_count += 1

            if saved_count > 0:
                messages.success(request, f"บันทึกข้อมูลสำเร็จ {saved_count} รายการ")
            if skipped_count > 0:
                messages.warning(request, f"ข้ามการบันทึกข้อมูลซ้ำ {skipped_count} รายการ")

            return redirect('report_list')

    else:
        # สร้างข้อมูลเริ่มต้นสำหรับฟอร์ม
        initial_data = [
            {
                'student_number': profile.student_number,
                'first_name': profile.first_name,
                'last_name': profile.last_name,
                'room': profile.room,
                'degree': profile.degree,
                'department': profile.department,
                'presence': False,
            }
            for profile in profiles
        ]
        formset = AttendanceCheckinFormSet(initial=initial_data, queryset=AttendanceCheckin.objects.none())

    # ดึงข้อมูลโปรไฟล์แล้วกรองค่า room และ department ที่ซ้ำกัน
    rooms_seen = set()
    departments_seen = set()
    grouped_results = []

    for profile in profiles:
        room = profile.room
        department = profile.department

        # หาก room ยังไม่เคยเจอและไม่เป็น None หรือว่าง ให้แสดงและเพิ่มใน rooms_seen
        if room and room not in rooms_seen:
            grouped_results.append({'room': room, 'department': None})
            rooms_seen.add(room)

        # หาก department ยังไม่เคยเจอและไม่เป็น None หรือว่าง ให้แสดงและเพิ่มใน departments_seen
        if department and department not in departments_seen:
            grouped_results.append({'room': None, 'department': department})
            departments_seen.add(department)

    # เรียงลำดับ grouped_results ตาม room และ department
    grouped_results.sort(key=lambda x: (x['room'] or "", x['department'] or ""))

    return render(request, 'attendance/bulk_checkin.html', {
        'formset': formset,
        'attendance': attendance,
        'grouped_profiles': grouped_results,  # ส่งผลลัพธ์ที่เรียงแล้วไปยัง template
        'room_filter': room_filter,
        'department_filter': department_filter
    })
class ReportList(ListView):
    model = Attendance
    template_name = 'attendance/report_list.html'
    context_object_name = 'reports'

def attendance_report(request, pk):
    # Fetch attendance record using the pk
    attendance = get_object_or_404(Attendance, pk=pk)
    attendances = Attendance.objects.all()
    progress_reports = {}
    rooms = set()  # Use a set to collect unique rooms
    departments = set()

    # Get filters from the request
    room_filter = request.GET.get('room', '').strip()
    department_filter = request.GET.get('department', '').strip()
    date_checkin = request.GET.get('date_checkin', '').strip()

    if request.method == 'GET':
        # Validate the date format
        if date_checkin and not parse_date(date_checkin):
            raise ValidationError("วันที่ไม่ถูกต้อง กรุณาใช้รูปแบบ YYYY-MM-DD")

        # Fetch attendance data and apply date filter
        attendance_data = AttendanceCheckin.objects.filter(att_name=attendance)
        if date_checkin:
            attendance_data = attendance_data.filter(date_checkin=date_checkin)

        # Apply room and department filters if selected
        if room_filter:
            attendance_data = attendance_data.filter(room=room_filter)
        if department_filter:
            attendance_data = attendance_data.filter(department=department_filter)

        attendance_count = {}
        for record in attendance_data:
            student_number = record.student_number
            presence_status = "Present" if record.presence else "Absent"

            if student_number not in attendance_count:
                attendance_count[student_number] = {
                    'present': 0,
                    'absent': 0,
                    'name': f"{record.first_name} {record.last_name}",
                    'att_name': attendance.att_name,
                    'room': record.room,
                    'department': record.department,
                }

            # Increment present or absent count
            if presence_status == "Present":
                attendance_count[student_number]['present'] += 1
            else:
                attendance_count[student_number]['absent'] += 1

            # Collect unique rooms and departments for filtering
            rooms.add(record.room)
            departments.add(record.department)

        # Prepare progress report with attendance percentage
        sorted_attendance_count = dict(sorted(attendance_count.items(), key=lambda item: (item[0] is None, item[0]))) # Sort by student_number
        for student_number, data in sorted_attendance_count.items():
            total_attendance = data['present'] + data['absent']
            attendance_percentage = (data['present'] / total_attendance * 100) if total_attendance > 0 else 0
            status = "ผ่าน" if attendance_percentage >= 60 else "ไม่ผ่าน"

            if data['att_name'] not in progress_reports:
                progress_reports[data['att_name']] = []

            progress_reports[data['att_name']].append({
                'student_number': student_number,
                'name': data['name'],
                'room': data['room'],
                'department': data['department'],
                'present': data['present'],
                'absent': data['absent'],
                'percentage': round(attendance_percentage, 2),
                'status': status,
            })

    # Remove duplicates and sort rooms and departments
    sorted_rooms = sorted(rooms, key=lambda x: (x.isdigit(), x))  # Unique rooms sorted
    sorted_departments = sorted(set(departments))  # Unique departments sorted
    

    context = {
        'attendances': attendances,
        'progress_reports': dict(sorted(progress_reports.items(), key=lambda item: (int(item[0]) if item[0].isdigit() else float('inf'), item[0]))),
        'rooms': sorted(set(sorted_rooms)),  # Ensure uniqueness before sorting
        'departments': sorted_departments,
        'room_filter': room_filter,
        'department_filter': department_filter,
        'date_checkin': date_checkin,
    }
    return render(request, 'attendance/attendance_report.html', context)

ACTIVITY_MAP = {
    'กิจกรรมเข้าแถว': 'line_up',
    'กิจกรรมชมรม': 'club',
    'กิจกรรมโฮมรูม': 'homeroom',
    'กิจกรรมลูกเสือ': 'scout'
}

@login_required
def sum_report(request):
    user_profile = get_object_or_404(Profile, user=request.user)
    student_number = user_profile.student_number

    all_profiles = Profile.objects.all()
    ticket_records = Ticket.objects.all()
    attendances = Attendance.objects.all()
    progress_reports = {}
    rooms = set()
    departments = set()

    # Get filters from the GET request
    room_filter = request.GET.get('room', '').strip()
    department_filter = request.GET.get('department', '').strip()
    term_filter = request.GET.get('term', '').strip()

    # 📌 คำนวณปีการศึกษาโดยใช้ปี พ.ศ. และเริ่มนับจากเดือนพฤษภาคม
    current_month = datetime.today().month
    current_year = datetime.today().year + 543  # แปลงเป็น พ.ศ.

    if current_month >= 5:
        academic_year_filter = str(current_year)
    else:
        academic_year_filter = str(current_year - 1)

    # รับค่าปีการศึกษาจาก URL ถ้ามี
    academic_year_filter = request.GET.get('academic_year', academic_year_filter).strip()

    report_form = ReportExportForm(request.GET)

    # Filter profiles based on room and department filters
    filtered_profiles = all_profiles
    if room_filter:
        filtered_profiles = filtered_profiles.filter(room__icontains=room_filter)
    if department_filter:
        filtered_profiles = filtered_profiles.filter(department__icontains=department_filter)

    # Populate the progress_reports dictionary for the filtered profiles
    for profile in filtered_profiles:
        student_number = profile.student_number

        if student_number not in progress_reports:
            progress_reports[student_number] = {
                'name': f"{profile.first_name} {profile.last_name}",
                'room': profile.room,
                'department': profile.department,
                'activities': {
                    'line_up': '-',
                    'club': '-',
                    'homeroom': '-',
                    'scout': '-',
                },
                'unit_count': 0,
                'status': '-',
                'overall_status': '-',
                'term': term_filter,
                'academic_year': academic_year_filter,
            }

        # Add room and department to the set (for all profiles)
        rooms.add(profile.room)
        departments.add(profile.department)

    # Process attendance data
    for attendance in attendances:
        attendance_data = AttendanceCheckin.objects.filter(att_name=attendance)

        # Filter attendance data based on the profile's room and department
        for record in attendance_data:
            student_number = record.student_number
            if student_number in progress_reports:
                presence_status = "ผ่าน" if record.presence else "ไม่ผ่าน"
                progress_reports[student_number]['activities'][ACTIVITY_MAP[attendance.att_name]] = presence_status
                rooms.add(record.room)
                departments.add(record.department)

    # Process ticket records
    ticket_summary = {}

    for ticket in ticket_records:
        student_number = ticket.student_number  

        if student_number not in ticket_summary:
            ticket_summary[student_number] = {
                'total_tickets': 0,
                'checked_in': 0,
                'total_units': 0,  
            }

        ticket_summary[student_number]['total_tickets'] += 1
        if ticket.checkin:
            ticket_summary[student_number]['checked_in'] += 1
            
            if ticket.activity.activity_category == '2 หน่วยกิจ':
                ticket_summary[student_number]['total_units'] += 2  
            elif ticket.activity.activity_category == '1 หน่วยกิจ':
                ticket_summary[student_number]['total_units'] += 1  

    # Update ticket data in the progress report
    for student_number, data in ticket_summary.items():
        total_tickets = data['total_tickets']
        checked_in_count = data['checked_in']
        total_units = data['total_units']

        if total_tickets > 0 and checked_in_count == total_tickets:
            status = "✅ยืนยันแล้ว"  
        else:
            status = "❌ยังไม่ยืนยัน"  
            
        data.update({
            'checked_in_percentage': (checked_in_count / total_tickets * 100) if total_tickets > 0 else 0,
            'status': status,
        })

        overall_status = "ไม่ผ่าน"  
        if total_units >= 6:
            overall_status = "ผ่าน"

        if student_number in progress_reports:
            progress_reports[student_number]['overall_status'] = overall_status  
            progress_reports[student_number]['unit_count'] = total_units  

    # Clean up progress_reports with valid entries
    valid_progress_reports = {
        student_number: report
        for student_number, report in progress_reports.items()
        if student_number is not None
    }

    # Ensure all rooms and departments are displayed
    rooms = {room for room in rooms if room is not None}
    departments = {department for department in departments if department is not None}

    context = {
        'progress_reports': dict(sorted(valid_progress_reports.items(), key=lambda item: (int(item[0]) if item[0].isdigit() else float('inf'), item[0]))),
        'rooms': sorted(rooms),
        'departments': sorted(departments),
        'room_filter': room_filter,
        'department_filter': department_filter,
        'term_filter': term_filter,  
        'academic_year_filter': academic_year_filter,  
        'report_form': report_form,  
    }

    return render(request, 'attendance/sum_report.html', context)

def export_to_excel(request):
    # ดึงข้อมูลการเข้าร่วมทั้งหมด
    attendances = Attendance.objects.all()
    progress_reports = {}

    # รับค่ากรองจาก request
    room_filter = request.GET.get('room', '').strip()
    department_filter = request.GET.get('department', '').strip()
    term_filter = request.GET.get('term', '').strip()  # รับค่ากรองเทอม
    academic_year_filter = request.GET.get('academic_year', '').strip()  # รับค่ากรองปีการศึกษา

    # วนลูปเพื่อดึงข้อมูลการเข้าร่วมแต่ละกิจกรรม
    for attendance in attendances:
        attendance_data = AttendanceCheckin.objects.filter(att_name=attendance)

        # กรองห้องและแผนกถ้ามี
        if room_filter:
            attendance_data = attendance_data.filter(room=room_filter)
        if department_filter:
            attendance_data = attendance_data.filter(department=department_filter)

        # เก็บข้อมูลนักเรียนแต่ละคนและสถานะกิจกรรม
        for record in attendance_data:
            student_number = record.student_number

            # ถ้ายังไม่มีข้อมูลนักเรียนใน progress_reports ให้เพิ่ม
            if student_number not in progress_reports:
                progress_reports[student_number] = {
                    'name': f"{record.first_name} {record.last_name}",
                    'room': record.room,
                    'department': record.department,
                    'activities': {
                        'line_up': '-',
                        'club': '-',
                        'homeroom': '-',
                        'scout': '-',
                    },
                    'overall_status': '-',
                    'term': term_filter,  # เพิ่มเทอม
                    'academic_year': academic_year_filter,  # เพิ่มปีการศึกษา
                }

            # อัปเดตสถานะกิจกรรมใน activities ของนักเรียน
            presence_status = "ผ" if record.presence else "มผ"
            progress_reports[student_number]['activities'][ACTIVITY_MAP[attendance.att_name]] = presence_status

    # สรุปข้อมูลตั๋วและสถานะโดยรวม
    for student_number, report in progress_reports.items():
        ticket_records = Ticket.objects.filter(student_number=student_number)
        total_units = 0

        if not ticket_records.exists():
            progress_reports[student_number]['overall_status'] = "-"
        else:
            for ticket in ticket_records:
                if ticket.checkin:
                    if ticket.activity.activity_category == '2 หน่วยกิจ':
                        total_units += 2
                    elif ticket.activity.activity_category == '1 หน่วยกิจ':
                        total_units += 1

            if total_units >= 6:
                progress_reports[student_number]['overall_status'] = "ผ"
            else:
                progress_reports[student_number]['overall_status'] = "มผ"

    # สร้าง workbook ใหม่
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'รายงานความก้าวหน้า'

    # ปรับขนาดความกว้างของคอลัมน์ (A ถึง I)
    worksheet.column_dimensions['A'].width = 20
    worksheet.column_dimensions['B'].width = 30
    worksheet.column_dimensions['C'].width = 30
    worksheet.column_dimensions['D'].width = 15
    worksheet.column_dimensions['E'].width = 15
    worksheet.column_dimensions['F'].width = 15
    worksheet.column_dimensions['G'].width = 15
    worksheet.column_dimensions['H'].width = 15
    worksheet.column_dimensions['I'].width = 20  # เพิ่มคอลัมน์สำหรับ Term และ Academic Year
    worksheet.column_dimensions['J'].width = 20

    # เพิ่มสีและจัดเส้นขอบให้หัวตาราง พร้อมฟอนต์
    header_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_font = Font(name='Angsana New', size=16, bold=True)  # ใช้ฟอนต์ Angsana New สำหรับหัวตาราง
    cell_font = Font(name='Angsana New', size=16)  # ใช้ฟอนต์ Angsana New สำหรับข้อมูล

    headers = ['รหัสประจำตัว', 'ชื่อ-สกุล', 'แผนก/ชั้น/กลุ่ม', 'กิจกรรมเข้าแถว', 'กิจกรรมชมรม', 'กิจกรรมโฮมรูม', 'กิจกรรมพิเศษ', 'กิจกรรมลูกเสือ', 'เทอม', 'ปีการศึกษา']
    worksheet.append(headers)

    for col in worksheet[1]:
        col.fill = header_fill
        col.border = border_style
        col.font = header_font
        col.alignment = Alignment(horizontal="center", vertical="center")  # จัดกึ่งกลางหัวตาราง

    # จัดเรียง student_number จากน้อยไปมาก
    sorted_progress_reports = sorted(progress_reports.items(), key=lambda item: int(item[0]) if item[0].isdigit() else float('inf'))

    # เขียนข้อมูลลงใน worksheet
    for student_number, details in sorted_progress_reports:
        row_data = [
            student_number,
            details['name'],
            f"{details['room']} {details['department']}",
            details['activities']['line_up'],
            details['activities']['club'],
            details['activities']['homeroom'],
            details['overall_status'],
            details['activities']['scout'],
            details['term'],  # เพิ่มข้อมูลเทอม
            details['academic_year'],  # เพิ่มข้อมูลปีการศึกษา
        ]
        worksheet.append(row_data)

        # เพิ่มเส้นขอบและฟอนต์ให้กับแต่ละเซลล์ในแถวข้อมูล
        for cell in worksheet[worksheet.max_row]:
            cell.border = border_style
            cell.font = cell_font
            cell.alignment = Alignment(horizontal="center", vertical="center")  # จัดกึ่งกลางข้อมูลในแถว

    # สร้าง HttpResponse สำหรับส่งออกไฟล์ Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="progress_report.xlsx"'
    workbook.save(response)

    return response

def daily_report(request):
    # รับค่าตัวกรองจาก GET request
    room_filter = request.GET.get('room', '').strip()
    department_filter = request.GET.get('department', '').strip()
    att_name_filter = request.GET.get('att_name', '').strip()
    date_filter = request.GET.get('date_checkin', '').strip()

    # ตรวจสอบว่ามีตัวกรองใดถูกเลือกหรือไม่
    filters_selected = any([room_filter, department_filter, att_name_filter, date_filter])

    # เรียกข้อมูลเฉพาะเมื่อมีตัวกรอง
    attendance_data = AttendanceCheckin.objects.none()  # เริ่มต้นด้วย QuerySet ว่าง
    if filters_selected:
        attendance_data = AttendanceCheckin.objects.all()
        if room_filter:
            attendance_data = attendance_data.filter(room=room_filter)
        if department_filter:
            attendance_data = attendance_data.filter(department=department_filter)
        if att_name_filter:
            attendance_data = attendance_data.filter(att_name__att_name=att_name_filter)
        if date_filter:
            try:
                date_checkin = datetime.strptime(date_filter, '%Y-%m-%d').date()
                attendance_data = attendance_data.filter(date_checkin=date_checkin)
            except ValueError:
                messages.error(request, "รูปแบบวันที่ไม่ถูกต้อง โปรดใช้ YYYY-MM-DD.")

        # เรียงตาม student_number จากน้อยไปมาก
        attendance_data = attendance_data.order_by('student_number')

    # ดึงค่าตัวเลือกสำหรับตัวกรอง
    rooms = AttendanceCheckin.objects.values_list('room', flat=True).distinct()
    departments = AttendanceCheckin.objects.values_list('department', flat=True).distinct()
    att_names = AttendanceCheckin.objects.values_list('att_name__att_name', flat=True).distinct()

    # Context ส่งไปที่ template
    context = {
        'attendance_data': attendance_data,
        'rooms': sorted(filter(None, rooms)),
        'departments': sorted(filter(None, departments)),
        'att_names': sorted(filter(None, att_names)),
        'room_filter': room_filter,
        'department_filter': department_filter,
        'att_name_filter': att_name_filter,
        'date_filter': date_filter,
        'filters_selected': filters_selected,  # ส่งตัวแปรไปที่ template
    }

    return render(request, 'attendance/daily_report.html', context)

@login_required
def self_report(request):
    # ดึงข้อมูลผู้ใช้
    user_profile = get_object_or_404(Profile, user=request.user)
    
    # ดึงบันทึกการเข้าร่วมสำหรับโปรไฟล์ของผู้ใช้
    attendance_records = AttendanceCheckin.objects.filter(student_number=user_profile.student_number)

    # ดึงบันทึกตั๋วสำหรับโปรไฟล์ของผู้ใช้
    ticket_records = Ticket.objects.filter(profile=user_profile)

    # พจนานุกรมสำหรับสรุปการเข้าร่วมตามกิจกรรม
    attendance_summary = {}
    
    # รวมข้อมูลการเข้าร่วมสำหรับแต่ละกิจกรรม
    for record in attendance_records:
        if record.att_name not in attendance_summary:
            attendance_summary[record.att_name] = {
                'present': 0,
                'absent': 0,
            }

        if record.presence:  # สมมติว่า presence เป็นฟิลด์บูลีน
            attendance_summary[record.att_name]['present'] += 1
        else:
            attendance_summary[record.att_name]['absent'] += 1

    # คำนวณเปอร์เซ็นต์และสถานะสำหรับแต่ละกิจกรรม
    for activity, data in attendance_summary.items():
        total = data['present'] + data['absent']
        attendance_percentage = (data['present'] / total * 100) if total > 0 else 0
        status = "ผ่าน" if attendance_percentage >= 60 else "ไม่ผ่าน"
        data.update({
            'attendance_percentage': attendance_percentage,
            'status': status,
            'activity': activity,
        })

    # สรุปข้อมูลตั๋ว
    ticket_summary = {}
    total_units = 0  # หน่วยรวมที่นับได้จากการเข้าร่วมกิจกรรมตั๋ว

    # รวมข้อมูลการเช็คอินตั๋ว
    for ticket in ticket_records:
        if ticket.activity not in ticket_summary:
            ticket_summary[ticket.activity] = {
                'total_tickets': 0,
                'checked_in': 0,
            }

        ticket_summary[ticket.activity]['total_tickets'] += 1
        if ticket.checkin:
            ticket_summary[ticket.activity]['checked_in'] += 1
            # เพิ่มหน่วยรวมตามหมวดหมู่กิจกรรม
            if ticket.activity.activity_category == '2 หน่วยกิจ':
                total_units += 2  # 2 หน่วยสำหรับกิจกรรมใหญ่
            elif ticket.activity.activity_category == '1 หน่วยกิจ':
                total_units += 1  # 1 หน่วยสำหรับกิจกรรมเล็ก

    # คำนวณสถานะการเช็คอินตั๋ว
    for activity, data in ticket_summary.items():
        total_tickets = data['total_tickets']
        checked_in_count = data['checked_in']
        
        # ตรวจสอบว่าผู้ใช้ได้เช็คอินตั๋วทั้งหมดหรือไม่
        if total_tickets > 0 and checked_in_count == total_tickets:
            status = "✅ยืนยันแล้ว"  # เช็คอินตั๋วทั้งหมดแล้ว
        else:
            status = "❌ยังไม่ยืนยัน"  # ยังไม่เช็คอินตั๋วทั้งหมด
            
        data.update({
            'checked_in_percentage': (checked_in_count / total_tickets * 100) if total_tickets > 0 else 0,
            'status': status,
            'activity': activity,
        })

    # กำหนดสถานะโดยรวมตามหน่วยกิจกรรมรวม
    overall_status = "ไม่ผ่าน"  # สถานะเริ่มต้น
    if total_units >= 6:
        overall_status = "ผ่าน"  # ผู้ใช้เข้าร่วมกิจกรรมเพียงพอ

    return render(request, 'attendance/self_report.html', {
        'attendance_summary': attendance_summary.values(),  # ส่งสรุปเป็นรายการ
        'ticket_summary': ticket_summary.values(),  # ส่งสรุปตั๋วเป็นรายการ
        'user_profile': user_profile,
        'total_units': total_units,  # ส่งหน่วยรวมสำหรับการแสดงผล
        'overall_status': overall_status,  # ส่งสถานะโดยรวม
    })

class SelectCheckin(ListView):
    model = Attendance
    template_name = 'attendance/select_checkin.html'
    context_object_name = 'attendances'

def retroactive_checkin(request, pk):
    attendance = get_object_or_404(Attendance, pk=pk)
    room_filter = request.GET.get('room', '').strip()
    department_filter = request.GET.get('department', '').strip()
    selected_date = request.GET.get('date_checkin')

    # ตรวจสอบว่าผู้ใช้เลือกวันที่หรือไม่
    if selected_date:
        try:
            selected_date = datetime.strptime(selected_date, "%Y-%m-%d").date()
        except ValueError:
            selected_date = None  # ถ้าเกิดข้อผิดพลาดให้มันเป็น None
    else:
        selected_date = None  # กรณีไม่ได้เลือกวันที่ใดๆ

    # Filter profiles and exclude null/empty rooms and departments
    profiles = Profile.objects.exclude(room__isnull=True).exclude(room='').exclude(department__isnull=True).exclude(department='')

    if room_filter:
        profiles = profiles.filter(room=room_filter)

    if department_filter:
        profiles = profiles.filter(department=department_filter)

    profiles = profiles.order_by('room', 'department')

    # Extract unique room and department values (distinct by room and department)
    unique_rooms = Profile.objects.values('room').distinct().order_by('room')
    unique_departments = Profile.objects.values('department').distinct().order_by('department')

    AttendanceCheckinFormSet = modelformset_factory(
        AttendanceCheckin,
        form=AttendanceCheckinForm,
        extra=len(profiles),
        can_delete=False
    )

    if request.method == 'POST':
        formset = AttendanceCheckinFormSet(request.POST)

        if formset.is_valid():
            instances = formset.save(commit=False)
            saved_count = 0
            skipped_count = 0

            for instance in instances:
                instance.att_name = attendance

                # กำหนด date_checkin เป็น selected_date ที่ผู้ใช้เลือก
                if selected_date:
                    instance.date_checkin = selected_date
                else:
                    instance.date_checkin = timezone.now().date()  # หากไม่ได้เลือกวันที่ให้ใช้วันที่ปัจจุบัน

                # ตรวจสอบข้อมูลซ้ำ
                exists_by_student = AttendanceCheckin.objects.filter(
                    student_number=instance.student_number,
                    att_name=attendance,
                    date_checkin=selected_date
                ).exists()

                exists_by_name = AttendanceCheckin.objects.filter(
                    first_name=instance.first_name,
                    last_name=instance.last_name,
                    att_name=attendance,
                    date_checkin=selected_date
                ).exists()

                if exists_by_student or exists_by_name:
                    skipped_count += 1
                    continue

                instance.save()
                saved_count += 1

            # แสดงข้อความการบันทึกข้อมูล
            if saved_count > 0:
                messages.success(request, f"บันทึกข้อมูลสำเร็จ {saved_count} รายการ")
            if skipped_count > 0:
                messages.warning(request, f"ข้ามการบันทึกข้อมูลซ้ำ {skipped_count} รายการ")

            return redirect('report_list')

    else:
        initial_data = [
            {
                'student_number': profile.student_number,
                'first_name': profile.first_name,
                'last_name': profile.last_name,
                'room': profile.room,
                'degree': profile.degree,
                'department': profile.department,
                'presence': False,
                'date_checkin': selected_date if selected_date else timezone.now().date()  # ใช้วันที่เลือกหรือวันที่ปัจจุบัน
            }
            for profile in profiles
        ]
        formset = AttendanceCheckinFormSet(initial=initial_data, queryset=AttendanceCheckin.objects.none())

    return render(request, 'attendance/retroactive_checkin.html', {
        'formset': formset,
        'attendance': attendance,
        'room_filter': room_filter,
        'department_filter': department_filter,
        'selected_date': selected_date,
        'unique_rooms': unique_rooms,  # unique room values for the filter
        'unique_departments': unique_departments,  # unique department values for the filter
    })

